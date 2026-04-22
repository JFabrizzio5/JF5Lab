"""Export a Excel (openpyxl). Devuelve bytes xlsx."""
from __future__ import annotations
import io
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment
from openpyxl.utils import get_column_letter


_HEADER_FONT = Font(bold=True, color="FFFFFF")
_HEADER_FILL = PatternFill("solid", fgColor="0F172A")


def rows_to_xlsx(sheet_name: str, headers: list[str], rows: list[list]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = sheet_name[:31] or "Sheet1"
    ws.append(headers)
    for i, _ in enumerate(headers, start=1):
        c = ws.cell(row=1, column=i)
        c.font = _HEADER_FONT
        c.fill = _HEADER_FILL
        c.alignment = Alignment(horizontal="center", vertical="center")
    for r in rows:
        ws.append(r)
    for i, _ in enumerate(headers, start=1):
        ws.column_dimensions[get_column_letter(i)].width = 20
    ws.freeze_panes = "A2"
    buf = io.BytesIO()
    wb.save(buf)
    return buf.getvalue()


def fmt_dt(dt: datetime | None) -> str:
    return dt.strftime("%Y-%m-%d %H:%M") if dt else ""
