"""PulsoMX v1 routes."""
from __future__ import annotations
import io
import os
from datetime import datetime, timedelta
from decimal import Decimal
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, HTMLResponse, RedirectResponse, JSONResponse
from sqlalchemy import select, func
from sqlalchemy.ext.asyncio import AsyncSession

from config.database import get_sql_db
from core.tenant import get_tenant_id, tenant_db
from modules.membership.models import (
    Tenant, Plan, TenantPlan, Venue, Room, Member, MembershipPlan, Membership,
    Instructor, ClassTemplate, ClassSession, Booking, CheckIn, Payment,
)
from modules.membership import schemas as S
from modules.membership.services.billing import (
    create_stripe_checkout, create_conekta_checkout,
    verify_stripe_webhook, verify_conekta_webhook, BillingError,
)

router = APIRouter()


# ─────── Plans (público) + Tenants ───────
@router.get("/plans", response_model=list[S.PlanOut])
async def list_saas_plans(db: AsyncSession = Depends(get_sql_db)):
    r = await db.execute(select(Plan).order_by(Plan.price_mxn))
    return r.scalars().all()


@router.post("/tenants", response_model=S.TenantOut, status_code=201)
async def create_tenant(payload: S.TenantCreate, db: AsyncSession = Depends(get_sql_db)):
    slug = payload.slug or payload.name.lower().replace(" ", "-")[:60]
    t = Tenant(name=payload.name, slug=slug, industry=payload.industry,
               rfc=payload.rfc, contact_email=payload.contact_email,
               contact_phone=payload.contact_phone)
    db.add(t)
    await db.flush()
    db.add(TenantPlan(tenant_id=t.id, plan_id=payload.plan_id or "free",
                     status="trialing", trial_end=datetime.utcnow() + timedelta(days=14),
                     current_period_end=datetime.utcnow() + timedelta(days=30)))
    await db.commit()
    await db.refresh(t)
    return t


# ─────── Venues ───────
@router.get("/venues", response_model=list[S.VenueOut])
async def list_venues(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Venue).order_by(Venue.name))
    return r.scalars().all()


@router.post("/venues", response_model=S.VenueOut, status_code=201)
async def create_venue(payload: S.VenueCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    v = Venue(tenant_id=tenant_id, **payload.model_dump())
    db.add(v); await db.commit(); await db.refresh(v)
    return v


# ─────── Members ───────
@router.get("/members", response_model=list[S.MemberOut])
async def list_members(q: str | None = None, db: AsyncSession = Depends(tenant_db)):
    query = select(Member).where(Member.active == True)  # noqa
    if q:
        like = f"%{q.lower()}%"
        query = query.where(
            func.lower(Member.first_name).like(like) |
            func.lower(Member.last_name).like(like) |
            Member.code.ilike(f"%{q}%") |
            Member.email.ilike(f"%{q}%")
        )
    r = await db.execute(query.order_by(Member.joined_at.desc()).limit(500))
    return r.scalars().all()


@router.post("/members", response_model=S.MemberOut, status_code=201)
async def create_member(payload: S.MemberCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    data = payload.model_dump()
    code = data.pop("code") or f"PM-{uuid4().hex[:6].upper()}"
    m = Member(tenant_id=tenant_id, code=code, **data)
    db.add(m); await db.commit(); await db.refresh(m)
    return m


@router.get("/members/{mid}", response_model=S.MemberOut)
async def get_member(mid: UUID, db: AsyncSession = Depends(tenant_db)):
    m = (await db.execute(select(Member).where(Member.id == mid))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Miembro no encontrado")
    return m


# ─────── Membership Plans (planes que ofrece el tenant) ───────
@router.get("/membership-plans", response_model=list[S.MembershipPlanOut])
async def list_mplans(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(MembershipPlan).where(MembershipPlan.active == True).order_by(MembershipPlan.price_mxn))
    return r.scalars().all()


@router.post("/membership-plans", response_model=S.MembershipPlanOut, status_code=201)
async def create_mplan(payload: S.MembershipPlanCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    p = MembershipPlan(tenant_id=tenant_id, **payload.model_dump())
    db.add(p); await db.commit(); await db.refresh(p)
    return p


@router.post("/memberships", response_model=S.MembershipOut, status_code=201)
async def create_membership(payload: S.MembershipCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    plan = (await db.execute(select(MembershipPlan).where(MembershipPlan.id == payload.membership_plan_id))).scalar_one_or_none()
    if not plan: raise HTTPException(404, "Plan no encontrado")
    expires = datetime.utcnow() + timedelta(days=plan.duration_days or 30)
    ms = Membership(tenant_id=tenant_id, member_id=payload.member_id,
                   membership_plan_id=plan.id, status="active",
                   expires_at=expires, visits_remaining=plan.visits_included,
                   auto_renew=payload.auto_renew)
    db.add(ms); await db.commit(); await db.refresh(ms)
    return ms


# ─────── Instructors ───────
@router.get("/instructors", response_model=list[S.InstructorOut])
async def list_instructors(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Instructor).where(Instructor.active == True).order_by(Instructor.name))
    return r.scalars().all()


@router.post("/instructors", response_model=S.InstructorOut, status_code=201)
async def create_instructor(payload: S.InstructorCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    i = Instructor(tenant_id=tenant_id, **payload.model_dump())
    db.add(i); await db.commit(); await db.refresh(i)
    return i


# ─────── Classes ───────
@router.get("/sessions", response_model=list[S.ClassSessionOut])
async def list_sessions(date_from: datetime | None = None, date_to: datetime | None = None, db: AsyncSession = Depends(tenant_db)):
    q = select(ClassSession)
    if date_from: q = q.where(ClassSession.starts_at >= date_from)
    if date_to:   q = q.where(ClassSession.starts_at <= date_to)
    r = await db.execute(q.order_by(ClassSession.starts_at).limit(500))
    return r.scalars().all()


@router.post("/sessions", response_model=S.ClassSessionOut, status_code=201)
async def create_session(payload: S.ClassSessionCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    s = ClassSession(tenant_id=tenant_id, **payload.model_dump())
    db.add(s); await db.commit(); await db.refresh(s)
    return s


@router.get("/class-templates", response_model=list[S.ClassTemplateOut])
async def list_templates(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(ClassTemplate).order_by(ClassTemplate.name))
    return r.scalars().all()


@router.post("/class-templates", response_model=S.ClassTemplateOut, status_code=201)
async def create_template(payload: S.ClassTemplateCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    t = ClassTemplate(tenant_id=tenant_id, **payload.model_dump())
    db.add(t); await db.commit(); await db.refresh(t)
    return t


# ─────── Bookings ───────
@router.post("/bookings", response_model=S.BookingOut, status_code=201)
async def create_booking(payload: S.BookingCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    sess = (await db.execute(select(ClassSession).where(ClassSession.id == payload.session_id))).scalar_one_or_none()
    if not sess: raise HTTPException(404, "Sesión no encontrada")
    if sess.booked_count >= sess.capacity: raise HTTPException(409, "Clase llena")
    b = Booking(tenant_id=tenant_id, session_id=payload.session_id, member_id=payload.member_id)
    sess.booked_count += 1
    db.add(b); await db.commit(); await db.refresh(b)
    return b


# ─────── Check-ins ───────
@router.post("/check-ins", response_model=S.CheckInOut)
async def checkin_punch(payload: S.CheckInPunch, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    code = payload.code
    if code.startswith("pulsomx://"): code = code.split("/")[-1]
    m = (await db.execute(select(Member).where(Member.code == code))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Miembro no encontrado para ese código")
    # verificar membresía vigente
    ms = (await db.execute(select(Membership).where(Membership.member_id == m.id, Membership.status == "active").order_by(Membership.started_at.desc()))).scalars().first()
    ci = CheckIn(tenant_id=tenant_id, member_id=m.id,
                venue_id=payload.venue_id, session_id=payload.session_id,
                method=payload.method, metadata_json={"has_active_membership": ms is not None})
    # descontar visita si aplica
    if ms and ms.visits_remaining is not None and ms.visits_remaining > 0:
        ms.visits_remaining -= 1
    db.add(ci); await db.commit(); await db.refresh(ci)
    return ci


@router.get("/check-ins", response_model=list[S.CheckInOut])
async def list_checkins(date_from: datetime | None = None, db: AsyncSession = Depends(tenant_db)):
    q = select(CheckIn)
    if date_from: q = q.where(CheckIn.at >= date_from)
    r = await db.execute(q.order_by(CheckIn.at.desc()).limit(500))
    return r.scalars().all()


# ─────── Member QR label ───────
@router.get("/members/{mid}/qr")
async def member_qr(mid: UUID, size: int = 320, db: AsyncSession = Depends(tenant_db)):
    m = (await db.execute(select(Member).where(Member.id == mid))).scalar_one_or_none()
    if not m: raise HTTPException(404, "Miembro")
    import qrcode
    from PIL import Image
    payload = f"pulsomx://{m.code}"
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(payload); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB").resize((size, size), Image.LANCZOS)
    buf = io.BytesIO(); img.save(buf, format="PNG")
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="image/png")


# ─────────────────────────────────────────────
# BILLING (Stripe + Conekta)
# ─────────────────────────────────────────────
@router.post("/billing/checkout", response_model=S.CheckoutResponse)
async def start_checkout(payload: S.CheckoutRequest, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    # Resolver monto + descripción
    amount = Decimal("0")
    description = ""
    if payload.plan_id:
        plan = (await db.execute(select(Plan).where(Plan.id == payload.plan_id))).scalar_one_or_none()
        if not plan: raise HTTPException(404, "SaaS plan no encontrado")
        amount = plan.price_mxn
        description = f"Suscripción PulsoMX — {plan.name}"
    elif payload.membership_plan_id:
        mp = (await db.execute(select(MembershipPlan).where(MembershipPlan.id == payload.membership_plan_id))).scalar_one_or_none()
        if not mp: raise HTTPException(404, "Membership plan")
        amount = mp.price_mxn
        description = mp.name
    else:
        raise HTTPException(400, "Falta plan_id o membership_plan_id")

    # Registrar Payment en estado pending
    reference = f"ref_{uuid4().hex[:16]}"
    pay = Payment(tenant_id=tenant_id, member_id=payload.member_id,
                 amount_mxn=amount, currency="MXN",
                 provider=payload.provider, provider_id=reference,
                 status="pending", description=description,
                 metadata_json={"plan_id": payload.plan_id, "mp_id": str(payload.membership_plan_id) if payload.membership_plan_id else None})
    db.add(pay); await db.commit(); await db.refresh(pay)

    try:
        if payload.provider == "conekta":
            url, sid = create_conekta_checkout(amount_mxn=amount, description=description, reference=reference)
        else:
            url, sid = create_stripe_checkout(amount_mxn=amount, description=description, reference=reference,
                                              success_url=payload.success_url, cancel_url=payload.cancel_url)
    except BillingError as e:
        raise HTTPException(502, str(e))

    pay.provider_id = sid or reference
    await db.commit()
    demo = not os.getenv("STRIPE_SECRET_KEY") if payload.provider == "stripe" else not os.getenv("CONEKTA_PRIVATE_KEY")
    return S.CheckoutResponse(provider=payload.provider, checkout_url=url, session_id=sid, demo=demo)


@router.get("/billing/demo-checkout/{provider}", response_class=HTMLResponse)
async def demo_checkout(provider: str, ref: str, amount: str):
    """Fallback: simula página de pago cuando no hay API keys."""
    brand = "#7c3aed" if provider == "stripe" else "#00b884"
    logo = "Stripe" if provider == "stripe" else "Conekta"
    html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Pagar con {logo} — PulsoMX</title>
<style>body{{margin:0;font-family:-apple-system,system-ui,sans-serif;background:linear-gradient(135deg,#faf7f2,#fff1e9);min-height:100vh;display:grid;place-items:center;color:#1a1527;}}
.card{{background:white;border-radius:20px;padding:40px;max-width:440px;width:90%;box-shadow:0 24px 48px rgba(0,0,0,.1);}}
h1{{font-size:26px;margin:0 0 6px;letter-spacing:-0.02em;}}
.brand{{color:{brand};font-weight:700;}}
.amount{{font-size:42px;font-weight:800;letter-spacing:-0.03em;margin:18px 0 4px;}}
.muted{{color:#6b5a72;font-size:14px;}}
input{{width:100%;padding:12px 14px;border:1.5px solid #e5d9cf;border-radius:10px;font-size:15px;margin:6px 0 14px;font-family:inherit;}}
button{{width:100%;padding:14px;border:none;background:{brand};color:white;font-weight:700;font-size:15px;border-radius:10px;cursor:pointer;}}
button:hover{{opacity:.9;}}
.badge{{display:inline-block;background:#fff7e0;color:#a16b00;padding:4px 10px;border-radius:999px;font-size:12px;font-weight:600;margin-bottom:16px;}}
</style></head><body><div class="card">
<div class="badge">🧪 Modo demo — sin API key</div>
<h1>Pagar con <span class="brand">{logo}</span></h1>
<div class="muted">Referencia: {ref}</div>
<div class="amount">${amount} <span style="font-size:14px;color:#6b5a72;font-weight:500;">MXN</span></div>
<div class="muted">Agrega tu <code>{provider.upper()}_SECRET_KEY</code> al .env para cobros reales.</div>
<form method="POST" action="/inventory/v1/__fake__" onsubmit="event.preventDefault(); fetch('/membership/v1/billing/demo-complete?ref={ref}',{{method:'POST'}}).then(()=>location.href='/billing/success?ref={ref}');">
<label style="font-size:13px;color:#6b5a72;">Tarjeta de prueba</label>
<input placeholder="4242 4242 4242 4242" value="4242 4242 4242 4242">
<div style="display:flex;gap:10px;">
<input placeholder="MM/AA" value="12/34" style="flex:1;">
<input placeholder="CVC" value="123" style="flex:1;">
</div>
<button type="submit">Pagar ${amount} MXN</button>
</form>
<div class="muted" style="margin-top:16px;text-align:center;font-size:12px;">No se cobra nada. Solo simula la respuesta.</div>
</div></body></html>"""
    return HTMLResponse(html)


@router.post("/billing/demo-complete")
async def demo_complete(ref: str, db: AsyncSession = Depends(get_sql_db)):
    p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
    if p:
        p.status = "succeeded"
        p.paid_at = datetime.utcnow()
        p.method = "card"
        await db.commit()
    return {"ok": True}


@router.post("/billing/webhook/stripe")
async def stripe_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    try:
        event = verify_stripe_webhook(payload, sig)
    except Exception as e:
        raise HTTPException(400, f"Firma inválida: {e}")
    if event.get("type") in ("checkout.session.completed", "payment_intent.succeeded"):
        obj = event["data"]["object"]
        ref = obj.get("metadata", {}).get("reference") or obj.get("client_reference_id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                p.status = "succeeded"
                p.paid_at = datetime.utcnow()
                p.receipt_url = obj.get("receipt_url")
                await db.commit()
    return {"received": True}


@router.post("/billing/webhook/conekta")
async def conekta_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    event = verify_conekta_webhook(payload, request.headers.get("digest"))
    if event.get("type", "").startswith("order.paid"):
        order = event["data"]["object"]
        ref = (order.get("metadata") or {}).get("reference") or order.get("id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                p.status = "succeeded"
                p.paid_at = datetime.utcnow()
                await db.commit()
    return {"received": True}


@router.get("/payments", response_model=list[S.PaymentOut])
async def list_payments(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Payment).order_by(Payment.created_at.desc()).limit(200))
    return r.scalars().all()


# ─────────────────────────────────────────────
# DASHBOARD + EXPORTS
# ─────────────────────────────────────────────
@router.get("/dashboard", response_model=S.DashboardKPIs)
async def dashboard(db: AsyncSession = Depends(tenant_db)):
    total_members = (await db.execute(select(func.count(Member.id)).where(Member.active == True))).scalar() or 0
    active_ms = (await db.execute(select(func.count(Membership.id)).where(Membership.status == "active"))).scalar() or 0
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    ci_today = (await db.execute(select(func.count(CheckIn.id)).where(CheckIn.at >= today_start))).scalar() or 0
    bookings_today = (await db.execute(select(func.count(Booking.id)).where(Booking.booked_at >= today_start))).scalar() or 0

    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    revenue = (await db.execute(
        select(func.coalesce(func.sum(Payment.amount_mxn), 0))
        .where(Payment.status == "succeeded", Payment.created_at >= month_start)
    )).scalar() or 0

    new_7 = (await db.execute(select(func.count(Member.id)).where(Member.joined_at >= datetime.utcnow() - timedelta(days=7)))).scalar() or 0
    churn = (await db.execute(select(func.count(Membership.id)).where(Membership.status == "canceled", Membership.started_at >= datetime.utcnow() - timedelta(days=30)))).scalar() or 0

    up_r = await db.execute(select(ClassSession).where(ClassSession.starts_at >= datetime.utcnow()).order_by(ClassSession.starts_at).limit(6))
    upcoming = [{"id": str(s.id), "name": s.name or "Clase", "starts_at": s.starts_at.isoformat(), "capacity": s.capacity, "booked": s.booked_count} for s in up_r.scalars().all()]

    ci_r = await db.execute(select(CheckIn, Member).join(Member, Member.id == CheckIn.member_id).order_by(CheckIn.at.desc()).limit(10))
    recent = [{"id": str(c.id), "member": f"{m.first_name} {m.last_name or ''}", "at": c.at.isoformat(), "method": c.method} for c, m in ci_r.all()]

    return S.DashboardKPIs(
        total_members=total_members, active_memberships=active_ms,
        check_ins_today=ci_today, bookings_today=bookings_today,
        revenue_mtd_mxn=Decimal(str(revenue)),
        churn_last_30d=churn, new_members_last_7d=new_7,
        upcoming_classes=upcoming, recent_check_ins=recent,
    )


@router.get("/exports/members.xlsx")
async def export_members(db: AsyncSession = Depends(tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    r = await db.execute(select(Member).order_by(Member.joined_at.desc()))
    wb = Workbook(); ws = wb.active; ws.title = "Miembros"
    hdr = ["Código", "Nombre", "Apellido", "Email", "Teléfono", "Ingresó", "Activo"]
    ws.append(hdr)
    for i, _ in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="7C3AED")
    for m in r.scalars().all():
        ws.append([m.code, m.first_name, m.last_name or "", m.email or "", m.phone or "",
                  m.joined_at.strftime("%Y-%m-%d %H:%M") if m.joined_at else "", "Sí" if m.active else "No"])
    buf = io.BytesIO(); wb.save(buf)
    return StreamingResponse(io.BytesIO(buf.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=miembros-{datetime.utcnow():%Y%m%d}.xlsx"})


@router.get("/exports/checkins.xlsx")
async def export_checkins(db: AsyncSession = Depends(tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    r = await db.execute(select(CheckIn, Member).join(Member, Member.id == CheckIn.member_id).order_by(CheckIn.at.desc()))
    wb = Workbook(); ws = wb.active; ws.title = "Check-ins"
    hdr = ["Fecha/Hora", "Código", "Miembro", "Tipo", "Método"]
    ws.append(hdr)
    for i, _ in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="7C3AED")
    for ci, m in r.all():
        ws.append([ci.at.strftime("%Y-%m-%d %H:%M"), m.code, f"{m.first_name} {m.last_name or ''}", ci.kind, ci.method])
    buf = io.BytesIO(); wb.save(buf)
    return StreamingResponse(io.BytesIO(buf.getvalue()),
                             media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
                             headers={"Content-Disposition": f"attachment; filename=checkins-{datetime.utcnow():%Y%m%d}.xlsx"})


# ─────── DEMO SEED ───────
@router.post("/demo/seed")
async def seed_demo(industry: str = "gym", db: AsyncSession = Depends(get_sql_db)):
    from modules.membership.services.demo import seed_for_industry
    return await seed_for_industry(db, industry)
