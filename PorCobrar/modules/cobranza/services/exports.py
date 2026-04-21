"""Excel exports con openpyxl."""
from __future__ import annotations
from io import BytesIO
from datetime import datetime
from openpyxl import Workbook
from openpyxl.styles import Font, PatternFill, Alignment


def _header(ws, headers: list[str]):
    for i, h in enumerate(headers, 1):
        c = ws.cell(row=1, column=i, value=h)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="10B981")
        c.alignment = Alignment(horizontal="center")


def _fit_cols(ws):
    for col in ws.columns:
        m = 10
        letter = col[0].column_letter
        for cell in col:
            try:
                v = len(str(cell.value or ""))
                if v > m:
                    m = v
            except Exception:
                pass
        ws.column_dimensions[letter].width = min(m + 2, 50)


def invoices_to_xlsx(invoices: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Facturas"
    headers = ["Folio", "Serie", "UUID CFDI", "Deudor", "RFC", "Emitida", "Vence", "Total", "Pagado", "Saldo", "Status"]
    _header(ws, headers)
    for inv in invoices:
        ws.append([
            inv.get("folio"), inv.get("serie"), inv.get("cfdi_uuid"), inv.get("debtor_name"),
            inv.get("debtor_rfc"),
            (inv.get("issued_at") or "").isoformat() if isinstance(inv.get("issued_at"), datetime) else inv.get("issued_at"),
            (inv.get("due_at") or "").isoformat() if isinstance(inv.get("due_at"), datetime) else inv.get("due_at"),
            float(inv.get("total") or 0),
            float(inv.get("paid_amount") or 0),
            float(inv.get("total") or 0) - float(inv.get("paid_amount") or 0),
            inv.get("status"),
        ])
    _fit_cols(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def debtors_to_xlsx(debtors: list[dict]) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Deudores"
    headers = ["Nombre", "RFC", "Email", "Teléfono", "Total adeudado", "Dias prom. atraso", "Score pago", "Ult. pago"]
    _header(ws, headers)
    for d in debtors:
        ws.append([
            d.get("name"), d.get("rfc"), d.get("email"), d.get("phone"),
            float(d.get("total_owed") or 0), int(d.get("overdue_days_avg") or 0),
            int(d.get("payment_score") or 0),
            (d.get("last_payment_at") or "").isoformat() if isinstance(d.get("last_payment_at"), datetime) else d.get("last_payment_at"),
        ])
    _fit_cols(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()


def aging_to_xlsx(buckets: dict) -> bytes:
    wb = Workbook()
    ws = wb.active
    ws.title = "Cartera Antiguedad"
    headers = ["Rango", "Facturas", "Monto"]
    _header(ws, headers)
    for label, data in buckets.items():
        ws.append([label, int(data.get("count") or 0), float(data.get("amount") or 0)])
    _fit_cols(ws)
    buf = BytesIO()
    wb.save(buf)
    return buf.getvalue()
