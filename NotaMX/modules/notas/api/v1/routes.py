"""NotaMX v1 API routes."""
from __future__ import annotations
import io
import os
from datetime import datetime, timedelta, date
from decimal import Decimal
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, HTTPException, Request
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession

from config.database import get_sql_db
from core.tenant import get_tenant_id, tenant_db
from modules.notas.models import (
    Tenant, Plan, TenantPlan, Customer, Product, Note, NoteItem,
    Payment, CfdiDocument, WhatsappTemplate, WhatsappMessage,
)
from modules.notas import schemas as S
from modules.notas.services.billing import (
    create_stripe_checkout, create_conekta_checkout,
    verify_stripe_webhook, verify_conekta_webhook, BillingError,
)
from modules.notas.services.whatsapp import send_text as wa_send

router = APIRouter()


# ═════════════════ PLANS (PUBLIC) + TENANTS ═════════════════
@router.get("/plans", response_model=list[S.PlanOut])
async def list_plans(db: AsyncSession = Depends(get_sql_db)):
    r = await db.execute(select(Plan).order_by(Plan.price_mxn))
    plans = r.scalars().all()
    if not plans:
        # bootstrap sin seed
        from modules.notas.services.demo import _ensure_plans
        await _ensure_plans(db)
        r = await db.execute(select(Plan).order_by(Plan.price_mxn))
        plans = r.scalars().all()
    return plans


@router.post("/tenants", response_model=S.TenantOut, status_code=201)
async def create_tenant(payload: S.TenantCreate, db: AsyncSession = Depends(get_sql_db)):
    slug = payload.slug or (payload.name.lower().replace(" ", "-")[:60] + f"-{uuid4().hex[:4]}")
    t = Tenant(
        name=payload.name, slug=slug, industry=payload.industry,
        rfc=payload.rfc, razon_social=payload.razon_social,
        regimen_fiscal=payload.regimen_fiscal, codigo_postal=payload.codigo_postal,
        contact_email=payload.contact_email, contact_phone=payload.contact_phone,
        whatsapp_number=payload.whatsapp_number,
    )
    db.add(t); await db.flush()
    db.add(TenantPlan(
        tenant_id=t.id, plan_id=payload.plan_id or "free",
        status="trialing", trial_end=datetime.utcnow() + timedelta(days=14),
        current_period_end=datetime.utcnow() + timedelta(days=30),
    ))
    await db.commit(); await db.refresh(t)
    return t


# ═════════════════ CUSTOMERS ═════════════════
@router.get("/customers", response_model=list[S.CustomerOut])
async def list_customers(q: str | None = None, db: AsyncSession = Depends(tenant_db)):
    query = select(Customer).where(Customer.active == True)
    if q:
        like = f"%{q.lower()}%"
        query = query.where(func.lower(Customer.name).like(like))
    r = await db.execute(query.order_by(Customer.created_at.desc()).limit(500))
    return r.scalars().all()


@router.post("/customers", response_model=S.CustomerOut, status_code=201)
async def create_customer(payload: S.CustomerCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    c = Customer(tenant_id=tenant_id, **payload.model_dump())
    db.add(c); await db.commit(); await db.refresh(c)
    return c


@router.get("/customers/{cid}", response_model=S.CustomerOut)
async def get_customer(cid: UUID, db: AsyncSession = Depends(tenant_db)):
    c = (await db.execute(select(Customer).where(Customer.id == cid))).scalar_one_or_none()
    if not c: raise HTTPException(404, "Cliente no encontrado")
    return c


# ═════════════════ PRODUCTS ═════════════════
@router.get("/products", response_model=list[S.ProductOut])
async def list_products(q: str | None = None, db: AsyncSession = Depends(tenant_db)):
    query = select(Product).where(Product.active == True)
    if q:
        like = f"%{q.lower()}%"
        query = query.where(func.lower(Product.name).like(like))
    r = await db.execute(query.order_by(Product.name).limit(500))
    return r.scalars().all()


@router.post("/products", response_model=S.ProductOut, status_code=201)
async def create_product(payload: S.ProductCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    p = Product(tenant_id=tenant_id, **payload.model_dump())
    db.add(p); await db.commit(); await db.refresh(p)
    return p


# ═════════════════ NOTES ═════════════════
def _recalculate(n: Note, items: list[NoteItem]):
    subtotal = Decimal("0"); tax_total = Decimal("0")
    for it in items:
        sub = (Decimal(str(it.qty)) * Decimal(str(it.unit_price))) - Decimal(str(it.discount or 0))
        tax = sub * Decimal(str(it.tax_rate or 0))
        it.subtotal = sub; it.tax = tax; it.total = sub + tax
        subtotal += sub; tax_total += tax
    n.subtotal = subtotal
    n.tax_total = tax_total
    n.total = subtotal + tax_total - Decimal(str(n.discount or 0))


async def _next_number(db: AsyncSession, tenant_id: str) -> str:
    count = (await db.execute(
        select(func.count(Note.id)).where(Note.tenant_id == tenant_id)
    )).scalar() or 0
    return f"NT-{count + 1:04d}"


@router.get("/notes", response_model=list[S.NoteOut])
async def list_notes(status: str | None = None, db: AsyncSession = Depends(tenant_db)):
    q = select(Note)
    if status:
        q = q.where(Note.status == status)
    r = await db.execute(q.order_by(Note.created_at.desc()).limit(500))
    return r.scalars().all()


@router.post("/notes", response_model=S.NoteDetailOut, status_code=201)
async def create_note(payload: S.NoteCreate, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    number = payload.number or await _next_number(db, tenant_id)
    n = Note(
        tenant_id=tenant_id, number=number, customer_id=payload.customer_id,
        status="draft", notes=payload.notes, valid_until=payload.valid_until,
        discount=payload.discount,
    )
    db.add(n); await db.flush()
    items: list[NoteItem] = []
    for i, it in enumerate(payload.items):
        ni = NoteItem(
            tenant_id=tenant_id, note_id=n.id, product_id=it.product_id,
            description=it.description, qty=it.qty, unit_price=it.unit_price,
            discount=it.discount or Decimal("0"), tax_rate=it.tax_rate,
            sort_order=i,
        )
        db.add(ni); items.append(ni)
    await db.flush()
    _recalculate(n, items)
    await db.commit(); await db.refresh(n)
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    out = S.NoteDetailOut.model_validate(n, from_attributes=True)
    out.items = [S.NoteItemOut.model_validate(i, from_attributes=True) for i in items]
    out.customer = S.CustomerOut.model_validate(cust, from_attributes=True) if cust else None
    return out


@router.get("/notes/{nid}", response_model=S.NoteDetailOut)
async def get_note(nid: UUID, db: AsyncSession = Depends(tenant_db)):
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    items = (await db.execute(select(NoteItem).where(NoteItem.note_id == nid).order_by(NoteItem.sort_order))).scalars().all()
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    out = S.NoteDetailOut.model_validate(n, from_attributes=True)
    out.items = [S.NoteItemOut.model_validate(i, from_attributes=True) for i in items]
    out.customer = S.CustomerOut.model_validate(cust, from_attributes=True) if cust else None
    return out


@router.post("/notes/{nid}/items", response_model=S.NoteItemOut, status_code=201)
async def add_note_item(nid: UUID, payload: S.NoteItemIn, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    ni = NoteItem(tenant_id=tenant_id, note_id=nid, **payload.model_dump())
    db.add(ni); await db.flush()
    items = (await db.execute(select(NoteItem).where(NoteItem.note_id == nid))).scalars().all()
    _recalculate(n, list(items))
    await db.commit(); await db.refresh(ni)
    return ni


@router.post("/notes/{nid}/send", response_model=S.NoteSendResponse)
async def send_note(nid: UUID, payload: S.NoteSendRequest, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    public_base = os.getenv("APP_PUBLIC_URL", "http://62.72.3.139:3050")
    public_url = f"{public_base}/n/{n.public_token}"

    sent_ok = False; status_msg = "skipped"
    if payload.channel == "whatsapp" and cust and (cust.whatsapp or cust.phone):
        body = payload.message or (
            f"Hola {cust.name}, te compartimos la nota {n.number} por ${float(n.total):,.2f} MXN. "
            f"Puedes verla y pagarla aquí: {public_url}"
        )
        sent_ok, status_msg = await wa_send(cust.whatsapp or cust.phone, body)
        db.add(WhatsappMessage(
            tenant_id=tenant_id, note_id=n.id, to_phone=(cust.whatsapp or cust.phone),
            body=body, status="sent" if sent_ok else "failed",
        ))
    n.status = "sent" if n.status == "draft" else n.status
    n.channel = payload.channel
    n.sent_at = datetime.utcnow()
    await db.commit()
    return S.NoteSendResponse(
        note_id=n.id, public_url=public_url, channel=payload.channel,
        whatsapp_sent=sent_ok, whatsapp_status=status_msg,
    )


# ═════════════════ PUBLIC NOTE VIEW ═════════════════
@router.get("/public/note/{token}", response_model=S.PublicNoteOut)
async def public_note(token: str, db: AsyncSession = Depends(get_sql_db)):
    # Sin RLS - buscamos con admin (get_sql_db sin tenant)
    n = (await db.execute(select(Note).where(Note.public_token == token))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    # configurar tenant post-lookup para filtrar hijos
    await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(n.tenant_id)))
    if not n.viewed_at:
        n.viewed_at = datetime.utcnow()
        await db.commit()
    items = (await db.execute(select(NoteItem).where(NoteItem.note_id == n.id).order_by(NoteItem.sort_order))).scalars().all()
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    tenant = (await db.execute(select(Tenant).where(Tenant.id == n.tenant_id))).scalar_one_or_none()
    return S.PublicNoteOut(
        id=n.id, number=n.number, status=n.status,
        subtotal=n.subtotal, tax_total=n.tax_total, discount=n.discount, total=n.total,
        currency=n.currency, valid_until=n.valid_until, notes=n.notes,
        items=[S.NoteItemOut.model_validate(i, from_attributes=True) for i in items],
        customer_name=cust.name if cust else None,
        tenant_name=tenant.name if tenant else "",
        tenant_brand_color=tenant.brand_color if tenant else "#10b981",
        tenant_rfc=tenant.rfc if tenant else None,
        paid_at=n.paid_at,
    )


# ═════════════════ CHECKOUT ═════════════════
@router.post("/notes/{nid}/checkout", response_model=S.CheckoutResponse)
async def note_checkout(nid: UUID, payload: S.CheckoutRequest, db: AsyncSession = Depends(get_sql_db)):
    # Puede llamarse con tenant (privado) o con token público; aquí usamos get_sql_db admin
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(n.tenant_id)))
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    reference = f"note_{n.public_token[:12]}"
    description = f"Nota {n.number or 'NT'} - {n.total} MXN"
    p = Payment(
        tenant_id=n.tenant_id, note_id=n.id, customer_id=n.customer_id,
        amount_mxn=n.total, provider=payload.provider, provider_id=reference,
        status="pending", description=description,
    )
    db.add(p); await db.commit(); await db.refresh(p)
    try:
        if payload.provider == "conekta":
            url, sid = create_conekta_checkout(
                amount_mxn=n.total, description=description, reference=reference,
                customer_name=cust.name if cust else None,
                customer_email=cust.email if cust else None,
                customer_phone=cust.phone if cust else None,
            )
        else:
            url, sid = create_stripe_checkout(
                amount_mxn=n.total, description=description, reference=reference,
                customer_email=cust.email if cust else None,
                success_url=payload.success_url, cancel_url=payload.cancel_url,
            )
    except BillingError as e:
        raise HTTPException(502, str(e))
    p.provider_id = sid or reference
    await db.commit()
    demo = not os.getenv("STRIPE_SECRET_KEY") if payload.provider == "stripe" else not os.getenv("CONEKTA_PRIVATE_KEY")
    return S.CheckoutResponse(provider=payload.provider, checkout_url=url, session_id=sid, demo=demo)


@router.get("/billing/demo-checkout/{provider}", response_class=HTMLResponse)
async def demo_checkout_page(provider: str, ref: str, amount: str):
    brand = "#635bff" if provider == "stripe" else "#00b884"
    logo = "Stripe" if provider == "stripe" else "Conekta"
    html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Pagar con {logo} — NotaMX</title>
<style>body{{margin:0;font-family:-apple-system,system-ui,sans-serif;background:linear-gradient(135deg,#f0fdf4,#ecfeff);min-height:100vh;display:grid;place-items:center;color:#111827;}}
.card{{background:white;border-radius:20px;padding:40px;max-width:440px;width:90%;box-shadow:0 24px 48px rgba(0,0,0,.1);}}
h1{{font-size:26px;margin:0 0 6px;letter-spacing:-0.02em;}}
.brand{{color:{brand};font-weight:700;}}
.amount{{font-size:42px;font-weight:800;letter-spacing:-0.03em;margin:18px 0 4px;}}
.muted{{color:#6b7280;font-size:14px;}}
input{{width:100%;padding:12px 14px;border:1.5px solid #e5e7eb;border-radius:10px;font-size:15px;margin:6px 0 14px;font-family:inherit;}}
button{{width:100%;padding:14px;border:none;background:{brand};color:white;font-weight:700;font-size:15px;border-radius:10px;cursor:pointer;}}
button:hover{{opacity:.9;}}
.badge{{display:inline-block;background:#fef9c3;color:#854d0e;padding:4px 10px;border-radius:999px;font-size:12px;font-weight:600;margin-bottom:16px;}}
.ok{{background:#d1fae5;color:#065f46;}}
</style></head><body><div class="card">
<div class="badge">Modo demo sin API key</div>
<h1>Pagar con <span class="brand">{logo}</span></h1>
<div class="muted">Referencia: {ref}</div>
<div class="amount">${amount} <span style="font-size:14px;color:#6b7280;font-weight:500;">MXN</span></div>
<div class="muted">Agrega <code>{provider.upper()}_SECRET_KEY</code> al .env para cobros reales.</div>
<form onsubmit="event.preventDefault(); fetch('/notas/v1/billing/demo-complete?ref={ref}',{{method:'POST'}}).then(()=>{{document.querySelector('.card').innerHTML='<div class=\\'badge ok\\'>Pago aprobado</div><h1>Gracias por tu pago</h1><div class=\\'muted\\'>Ya puedes cerrar esta ventana.</div>'; setTimeout(()=>window.close(),1500);}});">
<label style="font-size:13px;color:#6b7280;">Tarjeta de prueba</label>
<input placeholder="4242 4242 4242 4242" value="4242 4242 4242 4242">
<div style="display:flex;gap:10px;">
<input placeholder="MM/AA" value="12/34" style="flex:1;">
<input placeholder="CVC" value="123" style="flex:1;">
</div>
<button type="submit">Pagar ${amount} MXN</button>
</form>
<div class="muted" style="margin-top:16px;text-align:center;font-size:12px;">No se cobra nada. Solo simula la respuesta.</div>
</div></body></html>"""
    return HTMLResponse(html)


@router.post("/billing/demo-complete")
async def demo_complete(ref: str, db: AsyncSession = Depends(get_sql_db)):
    p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
    if p:
        await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(p.tenant_id)))
        p.status = "succeeded"
        p.paid_at = datetime.utcnow()
        p.method = "card"
        if p.note_id:
            n = (await db.execute(select(Note).where(Note.id == p.note_id))).scalar_one_or_none()
            if n:
                n.status = "paid"
                n.paid_at = datetime.utcnow()
        await db.commit()
    return {"ok": True}


@router.post("/billing/webhook/stripe")
async def stripe_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    try:
        event = verify_stripe_webhook(payload, sig)
    except Exception as e:
        raise HTTPException(400, f"Firma invalida: {e}")
    if event.get("type") in ("checkout.session.completed", "payment_intent.succeeded"):
        obj = event["data"]["object"]
        ref = obj.get("metadata", {}).get("reference") or obj.get("client_reference_id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(p.tenant_id)))
                p.status = "succeeded"; p.paid_at = datetime.utcnow()
                p.receipt_url = obj.get("receipt_url")
                if p.note_id:
                    n = (await db.execute(select(Note).where(Note.id == p.note_id))).scalar_one_or_none()
                    if n:
                        n.status = "paid"; n.paid_at = datetime.utcnow()
                await db.commit()
    return {"received": True}


@router.post("/billing/webhook/conekta")
async def conekta_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    event = verify_conekta_webhook(payload, request.headers.get("digest"))
    if event.get("type", "").startswith("order.paid"):
        order = event["data"]["object"]
        ref = (order.get("metadata") or {}).get("reference") or order.get("id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(p.tenant_id)))
                p.status = "succeeded"; p.paid_at = datetime.utcnow()
                if p.note_id:
                    n = (await db.execute(select(Note).where(Note.id == p.note_id))).scalar_one_or_none()
                    if n:
                        n.status = "paid"; n.paid_at = datetime.utcnow()
                await db.commit()
    return {"received": True}


@router.get("/payments", response_model=list[S.PaymentOut])
async def list_payments(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Payment).order_by(Payment.created_at.desc()).limit(200))
    return r.scalars().all()


# ═════════════════ CFDI (STUB) ═════════════════
@router.post("/notes/{nid}/cfdi", response_model=S.CfdiOut)
async def issue_cfdi(nid: UUID, payload: S.CfdiIssueRequest, tenant_id: str = Depends(get_tenant_id), db: AsyncSession = Depends(tenant_db)):
    """Stub CFDI 4.0. Marca pending y devuelve el documento.

    TODO producción: integrar con PAC (Finkok, Facturama, SW) para timbrado real:
        1. Generar XML CFDI 4.0 con datos tenant+customer+items
        2. Sellar con certificado CSD del tenant
        3. Timbrar vía PAC -> obtener UUID SAT + XML + PDF
        4. Guardar xml_url y pdf_url en storage
    """
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota no encontrada")
    if n.status != "paid":
        raise HTTPException(409, "Solo se puede timbrar una nota pagada")
    cust = (await db.execute(select(Customer).where(Customer.id == n.customer_id))).scalar_one_or_none() if n.customer_id else None
    pay = (await db.execute(select(Payment).where(Payment.note_id == n.id, Payment.status == "succeeded").order_by(Payment.paid_at.desc()))).scalars().first()
    # folio: count existentes + 1
    count = (await db.execute(select(func.count(CfdiDocument.id)))).scalar() or 0
    cfdi = CfdiDocument(
        tenant_id=tenant_id, related_note_id=n.id, payment_id=pay.id if pay else None,
        customer_id=n.customer_id, serie=payload.serie, folio=str(1000 + count),
        status="pending", total=n.total,
        uso_cfdi=payload.uso_cfdi or (cust.uso_cfdi if cust else "G03"),
        metodo_pago=payload.metodo_pago, forma_pago=payload.forma_pago,
    )
    db.add(cfdi); await db.commit(); await db.refresh(cfdi)
    return cfdi


@router.get("/cfdi", response_model=list[S.CfdiOut])
async def list_cfdi(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(CfdiDocument).order_by(CfdiDocument.created_at.desc()).limit(200))
    return r.scalars().all()


# ═════════════════ DASHBOARD ═════════════════
@router.get("/dashboard", response_model=S.DashboardKPIs)
async def dashboard(db: AsyncSession = Depends(tenant_db)):
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)
    notes_mtd = (await db.execute(select(func.count(Note.id)).where(Note.created_at >= month_start))).scalar() or 0
    notes_paid_mtd = (await db.execute(select(func.count(Note.id)).where(Note.status == "paid", Note.created_at >= month_start))).scalar() or 0
    notes_pending = (await db.execute(select(func.count(Note.id)).where(Note.status.in_(["draft", "sent"])))).scalar() or 0
    revenue = (await db.execute(
        select(func.coalesce(func.sum(Payment.amount_mxn), 0))
        .where(Payment.status == "succeeded", Payment.created_at >= month_start)
    )).scalar() or 0
    avg_amt = (await db.execute(
        select(func.coalesce(func.avg(Note.total), 0)).where(Note.status != "draft")
    )).scalar() or 0
    cfdi_mtd = (await db.execute(select(func.count(CfdiDocument.id)).where(CfdiDocument.created_at >= month_start))).scalar() or 0

    conv = (notes_paid_mtd / notes_mtd) if notes_mtd > 0 else 0.0

    # Top customers (by paid revenue)
    top_r = await db.execute(
        select(Customer.name, func.coalesce(func.sum(Payment.amount_mxn), 0).label("total"))
        .join(Payment, Payment.customer_id == Customer.id)
        .where(Payment.status == "succeeded")
        .group_by(Customer.id, Customer.name)
        .order_by(func.sum(Payment.amount_mxn).desc())
        .limit(5)
    )
    top = [{"name": name, "total_mxn": float(total)} for name, total in top_r.all()]

    # Recent notes
    rec_r = await db.execute(
        select(Note, Customer)
        .outerjoin(Customer, Customer.id == Note.customer_id)
        .order_by(Note.created_at.desc()).limit(10)
    )
    recent = [{
        "id": str(n.id), "number": n.number, "customer": c.name if c else "-",
        "total": float(n.total), "status": n.status,
        "created_at": n.created_at.isoformat(),
    } for n, c in rec_r.all()]

    # Monthly series (last 6 months)
    series = []
    now = datetime.utcnow()
    for i in range(5, -1, -1):
        m_start = (now.replace(day=1) - timedelta(days=i * 30)).replace(day=1, hour=0, minute=0, second=0, microsecond=0)
        m_end = (m_start + timedelta(days=32)).replace(day=1)
        s = (await db.execute(
            select(func.coalesce(func.sum(Payment.amount_mxn), 0))
            .where(Payment.status == "succeeded",
                   Payment.created_at >= m_start, Payment.created_at < m_end)
        )).scalar() or 0
        series.append({"month": m_start.strftime("%Y-%m"), "revenue": float(s)})

    return S.DashboardKPIs(
        notes_mtd=notes_mtd, notes_paid_mtd=notes_paid_mtd, notes_pending=notes_pending,
        revenue_mtd_mxn=Decimal(str(revenue)), avg_note_mxn=Decimal(str(avg_amt)),
        conversion_rate=round(conv, 3), cfdi_issued_mtd=cfdi_mtd,
        top_customers=top, recent_notes=recent, monthly_series=series,
    )


# ═════════════════ EXPORTS ═════════════════
@router.get("/exports/notes.xlsx")
async def export_notes(db: AsyncSession = Depends(tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    r = await db.execute(
        select(Note, Customer).outerjoin(Customer, Customer.id == Note.customer_id)
        .order_by(Note.created_at.desc())
    )
    wb = Workbook(); ws = wb.active; ws.title = "Notas"
    hdr = ["Folio", "Cliente", "Estado", "Canal", "Subtotal", "IVA", "Total", "Creada", "Pagada"]
    ws.append(hdr)
    for i, _ in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="10B981")
    for n, cust in r.all():
        ws.append([
            n.number or str(n.id)[:8], cust.name if cust else "-",
            n.status, n.channel,
            float(n.subtotal or 0), float(n.tax_total or 0), float(n.total or 0),
            n.created_at.strftime("%Y-%m-%d %H:%M") if n.created_at else "",
            n.paid_at.strftime("%Y-%m-%d %H:%M") if n.paid_at else "",
        ])
    buf = io.BytesIO(); wb.save(buf)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=notas-{datetime.utcnow():%Y%m%d}.xlsx"},
    )


@router.get("/exports/payments.xlsx")
async def export_payments(db: AsyncSession = Depends(tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    r = await db.execute(
        select(Payment, Customer).outerjoin(Customer, Customer.id == Payment.customer_id)
        .order_by(Payment.created_at.desc())
    )
    wb = Workbook(); ws = wb.active; ws.title = "Pagos"
    hdr = ["Cliente", "Monto", "Proveedor", "Método", "Estado", "Creado", "Pagado"]
    ws.append(hdr)
    for i, _ in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i); c.font = Font(bold=True, color="FFFFFF"); c.fill = PatternFill("solid", fgColor="10B981")
    for p, cust in r.all():
        ws.append([
            cust.name if cust else "-", float(p.amount_mxn or 0),
            p.provider or "", p.method or "", p.status,
            p.created_at.strftime("%Y-%m-%d %H:%M") if p.created_at else "",
            p.paid_at.strftime("%Y-%m-%d %H:%M") if p.paid_at else "",
        ])
    buf = io.BytesIO(); wb.save(buf)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=pagos-{datetime.utcnow():%Y%m%d}.xlsx"},
    )


# ═════════════════ QR PAGO ═════════════════
@router.get("/notes/{nid}/qr")
async def note_qr(nid: UUID, size: int = 320, db: AsyncSession = Depends(tenant_db)):
    n = (await db.execute(select(Note).where(Note.id == nid))).scalar_one_or_none()
    if not n: raise HTTPException(404, "Nota")
    import qrcode
    from PIL import Image
    public_base = os.getenv("APP_PUBLIC_URL", "http://62.72.3.139:3050")
    url = f"{public_base}/n/{n.public_token}"
    qr = qrcode.QRCode(error_correction=qrcode.constants.ERROR_CORRECT_M, box_size=10, border=2)
    qr.add_data(url); qr.make(fit=True)
    img = qr.make_image(fill_color="black", back_color="white").convert("RGB").resize((size, size), Image.LANCZOS)
    buf = io.BytesIO(); img.save(buf, format="PNG")
    return StreamingResponse(io.BytesIO(buf.getvalue()), media_type="image/png")


# ═════════════════ DEMO SEED ═════════════════
@router.post("/demo/seed")
async def seed_demo(industry: str = "freelance", db: AsyncSession = Depends(get_sql_db)):
    from modules.notas.services.demo import seed_for_industry
    return await seed_for_industry(db, industry)
