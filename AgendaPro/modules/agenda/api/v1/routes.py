"""AgendaPro v1 routes."""
from __future__ import annotations
import io
import os
import logging
from datetime import datetime, timedelta, date as DateType
from decimal import Decimal
from uuid import UUID, uuid4

from fastapi import APIRouter, Depends, HTTPException, Query, Request
from fastapi.responses import StreamingResponse, HTMLResponse, JSONResponse
from sqlalchemy import select, func, text
from sqlalchemy.ext.asyncio import AsyncSession

from config.database import get_sql_db
from core.tenant import get_tenant_id, tenant_db
from modules.agenda.models import (
    Tenant, Plan, TenantPlan, Staff, Service, AvailabilityRule, AvailabilityException,
    Customer, Appointment, Payment, Reminder, Review,
)
from modules.agenda import schemas as S
from modules.agenda.services.billing import (
    create_stripe_checkout, create_conekta_checkout,
    verify_stripe_webhook, verify_conekta_webhook, BillingError,
)
from modules.agenda.services.availability import compute_slots, has_conflict

router = APIRouter()
log = logging.getLogger("agenda")


# ─── Plans (público) + Tenants ───
@router.get("/plans", response_model=list[S.PlanOut])
async def list_saas_plans(db: AsyncSession = Depends(get_sql_db)):
    r = await db.execute(select(Plan).order_by(Plan.price_mxn))
    return r.scalars().all()


@router.post("/tenants", response_model=S.TenantOut, status_code=201)
async def create_tenant(payload: S.TenantCreate, db: AsyncSession = Depends(get_sql_db)):
    slug = (payload.slug or payload.name.lower().replace(" ", "-"))[:60]
    existing = (await db.execute(select(Tenant).where(Tenant.slug == slug))).scalar_one_or_none()
    if existing:
        raise HTTPException(409, f"Slug '{slug}' ya existe")
    t = Tenant(
        name=payload.name, slug=slug, industry=payload.industry,
        city=payload.city, brand_color=payload.brand_color,
        tagline=payload.tagline,
        contact_email=payload.contact_email, contact_phone=payload.contact_phone,
    )
    db.add(t); await db.flush()
    db.add(TenantPlan(
        tenant_id=t.id, plan_id=payload.plan_id or "free", status="trialing",
        trial_end=datetime.utcnow() + timedelta(days=14),
        current_period_end=datetime.utcnow() + timedelta(days=30),
    ))
    await db.commit(); await db.refresh(t)
    return t


# ─── Staff ───
@router.get("/staff", response_model=list[S.StaffOut])
async def list_staff(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Staff).where(Staff.active == True).order_by(Staff.name))  # noqa
    return r.scalars().all()


@router.post("/staff", response_model=S.StaffOut, status_code=201)
async def create_staff(payload: S.StaffCreate, tenant_id: str = Depends(get_tenant_id),
                       db: AsyncSession = Depends(tenant_db)):
    s = Staff(tenant_id=tenant_id, **payload.model_dump())
    db.add(s); await db.commit(); await db.refresh(s)
    return s


@router.patch("/staff/{sid}", response_model=S.StaffOut)
async def update_staff(sid: UUID, payload: S.StaffCreate, db: AsyncSession = Depends(tenant_db)):
    s = (await db.execute(select(Staff).where(Staff.id == sid))).scalar_one_or_none()
    if not s: raise HTTPException(404, "Staff")
    for k, v in payload.model_dump().items():
        setattr(s, k, v)
    await db.commit(); await db.refresh(s)
    return s


# ─── Services ───
@router.get("/services", response_model=list[S.ServiceOut])
async def list_services(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Service).where(Service.active == True).order_by(Service.name))  # noqa
    return r.scalars().all()


@router.post("/services", response_model=S.ServiceOut, status_code=201)
async def create_service(payload: S.ServiceCreate, tenant_id: str = Depends(get_tenant_id),
                         db: AsyncSession = Depends(tenant_db)):
    data = payload.model_dump()
    sv = Service(tenant_id=tenant_id, **data)
    db.add(sv); await db.commit(); await db.refresh(sv)
    return sv


@router.patch("/services/{sid}", response_model=S.ServiceOut)
async def update_service(sid: UUID, payload: S.ServiceCreate, db: AsyncSession = Depends(tenant_db)):
    sv = (await db.execute(select(Service).where(Service.id == sid))).scalar_one_or_none()
    if not sv: raise HTTPException(404, "Service")
    for k, v in payload.model_dump().items():
        setattr(sv, k, v)
    await db.commit(); await db.refresh(sv)
    return sv


# ─── Availability rules ───
@router.get("/availability-rules", response_model=list[S.AvailabilityRuleOut])
async def list_rules(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(AvailabilityRule).order_by(AvailabilityRule.weekday))
    return r.scalars().all()


@router.post("/availability-rules", response_model=S.AvailabilityRuleOut, status_code=201)
async def create_rule(payload: S.AvailabilityRuleCreate, tenant_id: str = Depends(get_tenant_id),
                     db: AsyncSession = Depends(tenant_db)):
    r = AvailabilityRule(tenant_id=tenant_id, **payload.model_dump())
    db.add(r); await db.commit(); await db.refresh(r)
    return r


# ─── Availability exceptions ───
@router.get("/availability-exceptions", response_model=list[S.AvailabilityExceptionOut])
async def list_exceptions(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(AvailabilityException).order_by(AvailabilityException.date))
    return r.scalars().all()


@router.post("/availability-exceptions", response_model=S.AvailabilityExceptionOut, status_code=201)
async def create_exception(payload: S.AvailabilityExceptionCreate, tenant_id: str = Depends(get_tenant_id),
                          db: AsyncSession = Depends(tenant_db)):
    e = AvailabilityException(tenant_id=tenant_id, **payload.model_dump())
    db.add(e); await db.commit(); await db.refresh(e)
    return e


# ─── Availability slots ───
@router.get("/availability")
async def get_availability(
    service_id: UUID,
    date_from: DateType,
    date_to: DateType,
    staff_id: UUID | None = None,
    db: AsyncSession = Depends(tenant_db),
):
    return await compute_slots(db, service_id, date_from, date_to, staff_id)


# ─── Customers ───
@router.get("/customers", response_model=list[S.CustomerOut])
async def list_customers(q: str | None = None, db: AsyncSession = Depends(tenant_db)):
    query = select(Customer)
    if q:
        like = f"%{q.lower()}%"
        query = query.where(func.lower(Customer.name).like(like) | Customer.phone.ilike(f"%{q}%"))
    r = await db.execute(query.order_by(Customer.created_at.desc()).limit(500))
    return r.scalars().all()


@router.post("/customers", response_model=S.CustomerOut, status_code=201)
async def create_customer(payload: S.CustomerCreate, tenant_id: str = Depends(get_tenant_id),
                          db: AsyncSession = Depends(tenant_db)):
    c = Customer(tenant_id=tenant_id, **payload.model_dump())
    db.add(c); await db.commit(); await db.refresh(c)
    return c


# ─── Appointments ───
@router.get("/appointments", response_model=list[S.AppointmentOut])
async def list_appointments(
    date_from: datetime | None = None,
    date_to: datetime | None = None,
    status: str | None = None,
    db: AsyncSession = Depends(tenant_db),
):
    q = select(Appointment)
    if date_from: q = q.where(Appointment.starts_at >= date_from)
    if date_to:   q = q.where(Appointment.starts_at <= date_to)
    if status:    q = q.where(Appointment.status == status)
    r = await db.execute(q.order_by(Appointment.starts_at).limit(500))
    return r.scalars().all()


@router.post("/appointments", response_model=S.AppointmentOut, status_code=201)
async def create_appointment(payload: S.AppointmentCreate, tenant_id: str = Depends(get_tenant_id),
                             db: AsyncSession = Depends(tenant_db)):
    svc = (await db.execute(select(Service).where(Service.id == payload.service_id))).scalar_one_or_none()
    if not svc: raise HTTPException(404, "Service")
    ends_at = payload.starts_at + timedelta(minutes=svc.duration_minutes)
    if await has_conflict(db, payload.staff_id, payload.starts_at, ends_at):
        raise HTTPException(409, "Conflicto: el staff ya tiene cita en ese horario")
    a = Appointment(
        tenant_id=tenant_id, customer_id=payload.customer_id, staff_id=payload.staff_id,
        service_id=payload.service_id, starts_at=payload.starts_at, ends_at=ends_at,
        status="confirmed" if not svc.requires_prepay else "pending_payment",
        total=svc.price_mxn, deposit=svc.deposit_mxn, notes=payload.notes,
        source="manual",
    )
    db.add(a); await db.commit(); await db.refresh(a)
    return a


@router.post("/appointments/{aid}/checkout", response_model=S.CheckoutResponse)
async def appt_checkout(aid: UUID, provider: str = "stripe", tenant_id: str = Depends(get_tenant_id),
                         db: AsyncSession = Depends(tenant_db)):
    a = (await db.execute(select(Appointment).where(Appointment.id == aid))).scalar_one_or_none()
    if not a: raise HTTPException(404, "Appointment")
    svc = (await db.execute(select(Service).where(Service.id == a.service_id))).scalar_one_or_none()
    cust = (await db.execute(select(Customer).where(Customer.id == a.customer_id))).scalar_one_or_none()
    amount = a.deposit if a.deposit and a.deposit > 0 else a.total
    desc = f"{svc.name if svc else 'Cita'} · {a.starts_at:%d/%b %H:%M}"
    reference = f"appt_{aid.hex[:12]}_{uuid4().hex[:6]}"

    pay = Payment(
        tenant_id=tenant_id, appointment_id=a.id, customer_id=a.customer_id,
        amount_mxn=amount, currency="MXN", provider=provider, provider_id=reference,
        status="pending", description=desc,
    )
    db.add(pay); await db.commit(); await db.refresh(pay)

    try:
        if provider == "conekta":
            url, sid = create_conekta_checkout(
                amount_mxn=amount, description=desc, reference=reference,
                customer_name=cust.name if cust else None,
                customer_email=cust.email if cust else None,
                customer_phone=cust.phone if cust else None,
            )
        else:
            url, sid = create_stripe_checkout(
                amount_mxn=amount, description=desc, reference=reference,
                customer_email=cust.email if cust else None,
            )
    except BillingError as e:
        raise HTTPException(502, str(e))

    pay.provider_id = sid or reference
    await db.commit()
    demo = not os.getenv("STRIPE_SECRET_KEY") if provider == "stripe" else not os.getenv("CONEKTA_PRIVATE_KEY")
    return S.CheckoutResponse(provider=provider, checkout_url=url, session_id=sid, demo=demo)


@router.post("/appointments/{aid}/confirm", response_model=S.AppointmentOut)
async def confirm_appointment(aid: UUID, db: AsyncSession = Depends(tenant_db)):
    a = (await db.execute(select(Appointment).where(Appointment.id == aid))).scalar_one_or_none()
    if not a: raise HTTPException(404, "Appointment")
    a.status = "confirmed"
    a.deposit_paid = True
    await db.commit(); await db.refresh(a)
    return a


@router.post("/appointments/{aid}/cancel", response_model=S.AppointmentOut)
async def cancel_appointment(aid: UUID, reason: str = "", db: AsyncSession = Depends(tenant_db)):
    a = (await db.execute(select(Appointment).where(Appointment.id == aid))).scalar_one_or_none()
    if not a: raise HTTPException(404, "Appointment")
    a.status = "canceled"
    a.cancel_reason = reason or "Cancelada"
    # TODO: stripe refund — stub
    log.info(f"[cancel] appt={aid} reason={reason} (refund stub)")
    await db.commit(); await db.refresh(a)
    return a


@router.post("/appointments/{aid}/complete", response_model=S.AppointmentOut)
async def complete_appointment(aid: UUID, tenant_id: str = Depends(get_tenant_id),
                               db: AsyncSession = Depends(tenant_db)):
    a = (await db.execute(select(Appointment).where(Appointment.id == aid))).scalar_one_or_none()
    if not a: raise HTTPException(404, "Appointment")
    a.status = "completed"
    # sumar visita al customer
    if a.customer_id:
        cust = (await db.execute(select(Customer).where(Customer.id == a.customer_id))).scalar_one_or_none()
        if cust: cust.total_visits = (cust.total_visits or 0) + 1
    # programar reminder de review (en 2h)
    db.add(Reminder(
        tenant_id=tenant_id, appointment_id=a.id,
        channel="whatsapp", send_at=datetime.utcnow() + timedelta(hours=2),
        status="scheduled", payload={"type": "review_request"},
    ))
    await db.commit(); await db.refresh(a)
    return a


@router.post("/appointments/{aid}/remind")
async def remind_appointment(aid: UUID, hours_before: int = 24,
                             tenant_id: str = Depends(get_tenant_id),
                             db: AsyncSession = Depends(tenant_db)):
    a = (await db.execute(select(Appointment).where(Appointment.id == aid))).scalar_one_or_none()
    if not a: raise HTTPException(404, "Appointment")
    send_at = a.starts_at - timedelta(hours=hours_before)
    rem = Reminder(
        tenant_id=tenant_id, appointment_id=a.id,
        channel="whatsapp", send_at=send_at, status="scheduled",
        payload={"type": "appointment_reminder", "hours_before": hours_before},
    )
    db.add(rem); await db.commit()
    # TODO: WhatsApp Cloud API
    log.info(f"[remind] appt={aid} send_at={send_at.isoformat()} channel=whatsapp (stub)")
    return {"ok": True, "scheduled_at": send_at.isoformat(), "channel": "whatsapp", "stub": True}


# ─── Dashboard ───
@router.get("/dashboard", response_model=S.DashboardKPIs)
async def dashboard(db: AsyncSession = Depends(tenant_db)):
    today_start = datetime.utcnow().replace(hour=0, minute=0, second=0, microsecond=0)
    week_end = today_start + timedelta(days=7)
    month_start = datetime.utcnow().replace(day=1, hour=0, minute=0, second=0, microsecond=0)

    appts_today = (await db.execute(
        select(func.count(Appointment.id)).where(
            Appointment.starts_at >= today_start,
            Appointment.starts_at < today_start + timedelta(days=1),
            Appointment.status.in_(["confirmed", "pending_payment"]),
        )
    )).scalar() or 0
    appts_week = (await db.execute(
        select(func.count(Appointment.id)).where(
            Appointment.starts_at >= today_start,
            Appointment.starts_at < week_end,
            Appointment.status.in_(["confirmed", "pending_payment"]),
        )
    )).scalar() or 0

    revenue = (await db.execute(
        select(func.coalesce(func.sum(Payment.amount_mxn), 0))
        .where(Payment.status == "succeeded", Payment.created_at >= month_start)
    )).scalar() or 0

    total_appts_m = (await db.execute(
        select(func.count(Appointment.id)).where(Appointment.starts_at >= month_start)
    )).scalar() or 0
    no_shows_m = (await db.execute(
        select(func.count(Appointment.id)).where(
            Appointment.starts_at >= month_start, Appointment.status == "no_show"
        )
    )).scalar() or 0
    confirmed_m = (await db.execute(
        select(func.count(Appointment.id)).where(
            Appointment.starts_at >= month_start, Appointment.status.in_(["confirmed", "completed"])
        )
    )).scalar() or 0

    no_show_rate = (no_shows_m / total_appts_m * 100) if total_appts_m else 0.0
    confirmed_rate = (confirmed_m / total_appts_m * 100) if total_appts_m else 0.0

    # top staff
    staff_r = await db.execute(
        select(Staff.name, func.count(Appointment.id).label("cnt"))
        .join(Appointment, Appointment.staff_id == Staff.id)
        .where(Appointment.starts_at >= month_start)
        .group_by(Staff.name).order_by(func.count(Appointment.id).desc()).limit(5)
    )
    top_staff = [{"name": n, "count": c} for n, c in staff_r.all()]

    svc_r = await db.execute(
        select(Service.name, func.count(Appointment.id).label("cnt"))
        .join(Appointment, Appointment.service_id == Service.id)
        .where(Appointment.starts_at >= month_start)
        .group_by(Service.name).order_by(func.count(Appointment.id).desc()).limit(5)
    )
    top_services = [{"name": n, "count": c} for n, c in svc_r.all()]

    up_r = await db.execute(
        select(Appointment, Customer, Service, Staff)
        .join(Customer, Customer.id == Appointment.customer_id, isouter=True)
        .join(Service, Service.id == Appointment.service_id, isouter=True)
        .join(Staff, Staff.id == Appointment.staff_id, isouter=True)
        .where(Appointment.starts_at >= datetime.utcnow(),
               Appointment.status.in_(["confirmed", "pending_payment"]))
        .order_by(Appointment.starts_at).limit(8)
    )
    upcoming = [{
        "id": str(a.id),
        "starts_at": a.starts_at.isoformat(),
        "customer": c.name if c else "—",
        "service": sv.name if sv else "—",
        "staff": st.name if st else "—",
        "status": a.status,
    } for a, c, sv, st in up_r.all()]

    return S.DashboardKPIs(
        appts_today=appts_today, appts_week=appts_week,
        revenue_mtd_mxn=Decimal(str(revenue)),
        no_show_rate=round(no_show_rate, 2),
        confirmed_rate=round(confirmed_rate, 2),
        top_staff=top_staff, top_services=top_services,
        upcoming=upcoming,
    )


@router.get("/exports/appointments.xlsx")
async def export_appointments(db: AsyncSession = Depends(tenant_db)):
    from openpyxl import Workbook
    from openpyxl.styles import Font, PatternFill
    r = await db.execute(
        select(Appointment, Customer, Service, Staff)
        .join(Customer, Customer.id == Appointment.customer_id, isouter=True)
        .join(Service, Service.id == Appointment.service_id, isouter=True)
        .join(Staff, Staff.id == Appointment.staff_id, isouter=True)
        .order_by(Appointment.starts_at.desc())
    )
    wb = Workbook(); ws = wb.active; ws.title = "Citas"
    hdr = ["Inicio", "Cliente", "Teléfono", "Servicio", "Staff", "Estado", "Total", "Pagado"]
    ws.append(hdr)
    for i, _ in enumerate(hdr, 1):
        c = ws.cell(row=1, column=i)
        c.font = Font(bold=True, color="FFFFFF")
        c.fill = PatternFill("solid", fgColor="1E3A8A")
    for a, cust, svc, staff in r.all():
        ws.append([
            a.starts_at.strftime("%Y-%m-%d %H:%M"),
            cust.name if cust else "",
            cust.phone if cust else "",
            svc.name if svc else "",
            staff.name if staff else "",
            a.status,
            float(a.total or 0),
            "Sí" if a.deposit_paid else "No",
        ])
    buf = io.BytesIO(); wb.save(buf)
    return StreamingResponse(
        io.BytesIO(buf.getvalue()),
        media_type="application/vnd.openxmlformats-officedocument.spreadsheetml.sheet",
        headers={"Content-Disposition": f"attachment; filename=citas-{datetime.utcnow():%Y%m%d}.xlsx"},
    )


# ─── DEMO SEED ───
@router.post("/demo/seed")
async def seed_demo(industry: str = "barber", db: AsyncSession = Depends(get_sql_db)):
    from modules.agenda.services.demo import seed_for_industry
    return await seed_for_industry(db, industry)


# ─────────────────────────────────────────────
# PÚBLICO (sin auth, resuelve tenant por slug)
# ─────────────────────────────────────────────
async def _resolve_tenant_by_slug(db: AsyncSession, slug: str) -> Tenant:
    t = (await db.execute(select(Tenant).where(Tenant.slug == slug))).scalar_one_or_none()
    if not t or not t.active:
        raise HTTPException(404, f"Negocio '{slug}' no encontrado")
    # activar RLS para esta sesión
    await db.execute(text("SELECT set_config('app.tenant_id', :tid, false)").bindparams(tid=str(t.id)))
    return t


@router.get("/public/tenant/{slug}", response_model=S.PublicTenantView)
async def public_tenant(slug: str, db: AsyncSession = Depends(get_sql_db)):
    t = await _resolve_tenant_by_slug(db, slug)
    svcs = (await db.execute(select(Service).where(Service.tenant_id == t.id, Service.active == True).order_by(Service.name))).scalars().all()  # noqa
    staff = (await db.execute(select(Staff).where(Staff.tenant_id == t.id, Staff.active == True).order_by(Staff.name))).scalars().all()  # noqa
    return S.PublicTenantView(
        id=t.id, name=t.name, slug=t.slug, industry=t.industry,
        brand_color=t.brand_color, logo_url=t.logo_url,
        tagline=t.tagline, about=t.about, city=t.city,
        instagram=t.instagram, timezone=t.timezone,
        services=svcs, staff=staff,
    )


@router.get("/public/availability/{slug}")
async def public_availability(
    slug: str,
    service_id: UUID,
    date: DateType | None = None,
    date_from: DateType | None = None,
    date_to: DateType | None = None,
    staff_id: UUID | None = None,
    db: AsyncSession = Depends(get_sql_db),
):
    await _resolve_tenant_by_slug(db, slug)
    if date:
        df = dt = date
    else:
        df = date_from or DateType.today()
        dt = date_to or (df + timedelta(days=7))
    return await compute_slots(db, service_id, df, dt, staff_id)


@router.post("/public/book/{slug}", response_model=S.PublicBookResponse)
async def public_book(slug: str, payload: S.PublicBookRequest, db: AsyncSession = Depends(get_sql_db)):
    t = await _resolve_tenant_by_slug(db, slug)
    svc = (await db.execute(select(Service).where(Service.id == payload.service_id, Service.tenant_id == t.id))).scalar_one_or_none()
    if not svc: raise HTTPException(404, "Servicio no encontrado")
    staff = (await db.execute(select(Staff).where(Staff.id == payload.staff_id, Staff.tenant_id == t.id))).scalar_one_or_none()
    if not staff: raise HTTPException(404, "Staff no encontrado")

    ends_at = payload.starts_at + timedelta(minutes=svc.duration_minutes)
    if await has_conflict(db, staff.id, payload.starts_at, ends_at):
        raise HTTPException(409, "Ese horario ya fue tomado")

    # upsert customer por phone
    cust = (await db.execute(select(Customer).where(
        Customer.tenant_id == t.id, Customer.phone == payload.customer.phone
    ))).scalar_one_or_none()
    if not cust:
        cust = Customer(
            tenant_id=t.id, name=payload.customer.name,
            phone=payload.customer.phone, email=payload.customer.email,
            whatsapp_optin=True,
        )
        db.add(cust); await db.flush()
    else:
        cust.name = payload.customer.name
        if payload.customer.email: cust.email = payload.customer.email

    initial_status = "pending_payment" if svc.requires_prepay else "confirmed"
    a = Appointment(
        tenant_id=t.id, customer_id=cust.id, staff_id=staff.id, service_id=svc.id,
        starts_at=payload.starts_at, ends_at=ends_at, status=initial_status,
        total=svc.price_mxn, deposit=svc.deposit_mxn, notes=payload.notes,
        source="public",
    )
    db.add(a); await db.commit(); await db.refresh(a)

    checkout_url = None
    if svc.requires_prepay:
        amount = svc.deposit_mxn if svc.deposit_mxn and svc.deposit_mxn > 0 else svc.price_mxn
        reference = f"appt_{a.id.hex[:12]}_{uuid4().hex[:6]}"
        desc = f"{svc.name} · {payload.starts_at:%d/%b %H:%M} · {t.name}"
        pay = Payment(
            tenant_id=t.id, appointment_id=a.id, customer_id=cust.id,
            amount_mxn=amount, currency="MXN", provider="stripe",
            provider_id=reference, status="pending", description=desc,
        )
        db.add(pay); await db.commit()
        try:
            url, sid = create_stripe_checkout(
                amount_mxn=amount, description=desc, reference=reference,
                customer_email=cust.email,
            )
            checkout_url = url
            pay.provider_id = sid or reference
            await db.commit()
        except BillingError:
            checkout_url = None

    return S.PublicBookResponse(
        appointment_id=a.id, status=a.status,
        requires_payment=svc.requires_prepay,
        checkout_url=checkout_url,
        total=svc.price_mxn, deposit=svc.deposit_mxn,
    )


# ─────────────────────────────────────────────
# BILLING WEBHOOKS + DEMO CHECKOUT
# ─────────────────────────────────────────────
@router.get("/billing/demo-checkout/{provider}", response_class=HTMLResponse)
async def demo_checkout(provider: str, ref: str, amount: str):
    brand = "#635bff" if provider == "stripe" else "#00b884"
    logo = "Stripe" if provider == "stripe" else "Conekta"
    public = os.getenv("APP_PUBLIC_URL", "http://62.72.3.139:3055")
    html = f"""<!doctype html><html lang="es"><head><meta charset="utf-8"><title>Pagar — AgendaPro</title>
<style>
@import url('https://fonts.googleapis.com/css2?family=Playfair+Display:ital,wght@0,500;1,500&family=Inter:wght@400;600;700&display=swap');
body{{margin:0;font-family:'Inter',system-ui,sans-serif;background:linear-gradient(135deg,#fdfbf5,#f5efe2);min-height:100vh;display:grid;place-items:center;color:#0c1933;}}
.card{{background:white;border-radius:18px;padding:40px;max-width:460px;width:92%;box-shadow:0 24px 64px rgba(12,25,51,.12);border:1px solid #eadbc0;}}
h1{{font-family:'Playfair Display',serif;font-style:italic;font-size:30px;margin:0 0 6px;letter-spacing:-0.02em;}}
.brand{{color:{brand};font-weight:700;font-style:normal;}}
.amount{{font-size:44px;font-weight:700;letter-spacing:-0.03em;margin:18px 0 4px;color:#0c1933;}}
.muted{{color:#5b6780;font-size:13px;}}
input{{width:100%;padding:13px 15px;border:1.5px solid #e2d3b8;border-radius:10px;font-size:15px;margin:6px 0 14px;font-family:inherit;box-sizing:border-box;background:#fdfbf5;}}
button{{width:100%;padding:14px;border:none;background:#0c1933;color:#fdfbf5;font-weight:700;font-size:15px;border-radius:10px;cursor:pointer;letter-spacing:0.02em;}}
button:hover{{background:#1a2746;}}
.badge{{display:inline-block;background:#fdf2d4;color:#8a5a00;padding:4px 12px;border-radius:999px;font-size:12px;font-weight:700;margin-bottom:16px;letter-spacing:0.05em;}}
.accent{{color:#b6892b;font-weight:700;}}
</style></head><body><div class="card">
<div class="badge">MODO DEMO · SIN API KEY</div>
<h1>Pagar con <span class="brand">{logo}</span></h1>
<div class="muted">Ref: <span class="accent">{ref}</span></div>
<div class="amount">${amount} <span style="font-size:14px;color:#5b6780;font-weight:500;">MXN</span></div>
<div class="muted" style="margin-bottom:18px;">Agrega <code>{provider.upper()}_SECRET_KEY</code> al .env para cobros reales.</div>
<form onsubmit="event.preventDefault(); fetch('/agenda/v1/billing/demo-complete?ref={ref}',{{method:'POST'}}).then(()=>location.href='{public}/book/success?ref={ref}');">
<label style="font-size:12px;color:#5b6780;font-weight:600;letter-spacing:0.03em;">TARJETA DE PRUEBA</label>
<input placeholder="4242 4242 4242 4242" value="4242 4242 4242 4242">
<div style="display:flex;gap:10px;">
<input placeholder="MM/AA" value="12/34" style="flex:1;">
<input placeholder="CVC" value="123" style="flex:1;">
</div>
<button type="submit">Pagar ${amount} MXN</button>
</form>
<div class="muted" style="margin-top:16px;text-align:center;font-size:11px;">No se cobra nada. Solo simula la respuesta.</div>
</div></body></html>"""
    return HTMLResponse(html)


@router.post("/billing/demo-complete")
async def demo_complete(ref: str, db: AsyncSession = Depends(get_sql_db)):
    p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
    if p:
        p.status = "succeeded"
        p.paid_at = datetime.utcnow()
        p.method = "card"
        if p.appointment_id:
            a = (await db.execute(select(Appointment).where(Appointment.id == p.appointment_id))).scalar_one_or_none()
            if a:
                a.deposit_paid = True
                a.status = "confirmed"
        await db.commit()
    return {"ok": True}


@router.post("/billing/webhook/stripe")
async def stripe_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    sig = request.headers.get("stripe-signature", "")
    try:
        event = verify_stripe_webhook(payload, sig)
    except Exception as e:
        raise HTTPException(400, f"Firma inválida: {e}")
    if event.get("type") in ("checkout.session.completed", "payment_intent.succeeded"):
        obj = event["data"]["object"]
        ref = obj.get("metadata", {}).get("reference") or obj.get("client_reference_id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                p.status = "succeeded"
                p.paid_at = datetime.utcnow()
                if p.appointment_id:
                    a = (await db.execute(select(Appointment).where(Appointment.id == p.appointment_id))).scalar_one_or_none()
                    if a:
                        a.deposit_paid = True
                        a.status = "confirmed"
                await db.commit()
    return {"received": True}


@router.post("/billing/webhook/conekta")
async def conekta_webhook(request: Request, db: AsyncSession = Depends(get_sql_db)):
    payload = await request.body()
    event = verify_conekta_webhook(payload, request.headers.get("digest"))
    if event.get("type", "").startswith("order.paid"):
        order = event["data"]["object"]
        ref = (order.get("metadata") or {}).get("reference") or order.get("id")
        if ref:
            p = (await db.execute(select(Payment).where(Payment.provider_id == ref))).scalar_one_or_none()
            if p:
                p.status = "succeeded"
                p.paid_at = datetime.utcnow()
                if p.appointment_id:
                    a = (await db.execute(select(Appointment).where(Appointment.id == p.appointment_id))).scalar_one_or_none()
                    if a:
                        a.deposit_paid = True; a.status = "confirmed"
                await db.commit()
    return {"received": True}


@router.get("/payments", response_model=list[S.PaymentOut])
async def list_payments(db: AsyncSession = Depends(tenant_db)):
    r = await db.execute(select(Payment).order_by(Payment.created_at.desc()).limit(200))
    return r.scalars().all()
